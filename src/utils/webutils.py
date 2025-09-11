import os
import json
import requests
import time
import pandas as pd
import openpyxl
import logging
from bs4 import BeautifulSoup
from urllib.parse import urljoin

from config import JGI_API_BASE_URL, FILES_PER_PAGE, REQUEST_DELAY
from src.credentials import HEADERS, COOKIES


def download_mycocosm_fungi_table(url: str, local_xlsx: str) -> pd.DataFrame:
    """
    Downloads the HTML table from the given MycoCosm URL and returns it as a DataFrame.
    If the download fails, attempts to load and wrangle a local Excel file.
    If both fail, instructs the user to manually download the table.

    Args:
        url (str): The URL of the web page containing the table.
        local_xlsx (str): Path to a local Excel file to use as a fallback.

    Returns:
        pd.DataFrame: The table data as a pandas DataFrame, or None if unavailable.
    """
    try:
        response = requests.get(url, cookies=COOKIES, headers=HEADERS)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        table = soup.find('table')
        if not table:
            logging.error("No table found at the provided URL.")
            return None

        headers = [th.get_text(strip=True) for th in table.find_all('th')]
        rows = []
        for tr in table.find_all('tr')[1:]:
            cells = tr.find_all('td')
            row_data = []
            row_links = []
            for cell in cells:
                cell_text = cell.get_text(strip=True)
                row_data.append(cell_text)
                link_tag = cell.find('a')
                if link_tag and link_tag.get('href'):
                    full_url = urljoin(url, link_tag['href'])
                    row_links.append(full_url)
                else:
                    row_links.append(None)
            combined_row = row_data + row_links
            rows.append(combined_row)

        link_headers = [f"{col}_link" for col in headers]
        final_headers = headers + link_headers

        df = pd.DataFrame(rows, columns=final_headers)
        df['portal'] = df['Name_link'].str.replace('https://mycocosm.jgi.doe.gov/', '')
        df['reference'] = df['Published_link']
        df = df.loc[:, ~df.columns.str.endswith('_link')]

        logging.info(f"Successfully downloaded table from {url}")
        return df

    except Exception as e:
        logging.error(f"Error downloading table: {e}")
        # Try to load local Excel file
        if os.path.exists(local_xlsx):
            try:
                # Load with pandas for values
                df = pd.read_excel(local_xlsx)
                # Load with openpyxl for hyperlinks
                wb = openpyxl.load_workbook(local_xlsx, data_only=True)
                ws = wb.active
                # Find the column indices for "Name" and "Published"
                header_row = next(ws.iter_rows(min_row=1, max_row=1))
                name_col_idx = None
                published_col_idx = None
                for idx, cell in enumerate(header_row):
                    if cell.value == "Name":
                        name_col_idx = idx
                    if cell.value == "Published":
                        published_col_idx = idx
                # Extract Name_link and Published_link
                name_links = []
                published_links = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    # Name link
                    if name_col_idx is not None:
                        cell = row[name_col_idx]
                        if cell.hyperlink:
                            name_links.append(cell.hyperlink.target)
                        else:
                            name_links.append(None)
                    else:
                        name_links.append(None)
                    # Published link
                    if published_col_idx is not None:
                        cell = row[published_col_idx]
                        if cell.hyperlink:
                            published_links.append(cell.hyperlink.target)
                        else:
                            published_links.append(None)
                    else:
                        published_links.append(None)
                if name_col_idx is not None:
                    df['Name_link'] = name_links
                    df['portal'] = df['Name_link'].str.replace('https://mycocosm.jgi.doe.gov/', '', regex=False)
                else:
                    df['portal'] = df['Name'].astype(str).str.replace('https://mycocosm.jgi.doe.gov/', '', regex=False)
                if published_col_idx is not None:
                    df['Published_link'] = published_links
                    df['reference'] = df['Published_link']
                elif 'Published' in df.columns:
                    df['reference'] = df['Published']
                
                df = df.loc[:, ~df.columns.str.endswith('_link')]
                
                logging.info(f"Loaded and wrangled local Excel file: {local_xlsx}")
                return df
            except Exception as e2:
                logging.error(f"Failed to load or wrangle local Excel file: {e2}")
        # If both fail, instruct user to manually download
        print(
            "\n❌ Could not download the MycoCosm fungi table automatically and no valid local file was found.\n"
            "Please manually download the table as Excel from:\n"
            "    https://mycocosm.jgi.doe.gov/fungi/fungi.info.html\n"
            "and save it as:\n"
            f"    {local_xlsx}\n"
        )
        return None

def fetch_portal_json(organism_id: str, header: dict, out_dir: str) -> bool:
    """
    Fetch the json listing of all files for a given organism ID from JGI API.

    Args:
        organism_id (str): The ID of the organism to fetch files for.
    """
    logging.info(f"Fetching all files for {organism_id} from JGI...")
    params = {
        "organism": organism_id,
        "api_version": 2,
        "a": "false",
        "h": "false",
        "d": "asc",
        "p": 1,
        "x": FILES_PER_PAGE,
        "t": "simple"
    }

    os.makedirs(out_dir, exist_ok=True)
    page = 1
    total_files = 0

    try:
        while True:
            params["p"] = page
            logging.info(f"Fetching page {page} for {organism_id}...")
            response = requests.get(JGI_API_BASE_URL, params=params, headers=header)
            response.raise_for_status()
            data = response.json()

            current_files = data.get("organisms", [{}])[0].get("files", [])
            if not current_files:
                logging.info(f"No more files found for {organism_id}. Stopping pagination.")
                break

            total_files += len(current_files)
            page_filename = os.path.join(out_dir, f"all_files_{organism_id}_page_{page}.json")
            with open(page_filename, "w") as f:
                json.dump(data, f, indent=2)
            logging.info(f"Saved page {page} to {page_filename}")

            page += 1
            time.sleep(REQUEST_DELAY)

        if total_files == 0:
            logging.warning(f"No files found for {organism_id} across any pages.")
        else:
            logging.info(f"Total {total_files} files saved for {organism_id}.")
        return total_files > 0

    except requests.exceptions.HTTPError as e:
        logging.error(f"HTTP error for {organism_id}: {e} - {response.status_code}: {response.text}")
        return False
    except requests.exceptions.RequestException as e:
        logging.error(f"Request error for {organism_id}: {e}")
        return False
    except Exception as e:
        logging.error(f"General error for {organism_id}: {e}")
        return False

def batch_fetch_json(organism_ids: list, headers: dict, out_dir: str):
    """
    Batch fetch the json listing of all files for a given organism ID from JGI API.

    Args:
        organism_ids (list): List of organism IDs to fetch files for.
    """
    from concurrent.futures import ThreadPoolExecutor
    os.makedirs(out_dir, exist_ok=True)

    def fetch_if_needed(organism_id):
        if not any(f.startswith(f"all_files_{organism_id}_page_") for f in os.listdir(out_dir)):
            fetch_portal_json(organism_id, headers, out_dir)
        else:
            logging.info(f"Using cached JSON for {organism_id}...")

    with ThreadPoolExecutor(max_workers=5) as executor:
        executor.map(fetch_if_needed, organism_ids)

def parse_portal_jsons(organism_ids: list, input_dir: str) -> pd.DataFrame:
    """
    Parses the JSON files obtained from mycocosm and returns them as a DataFrame.

    Args:
        organism_ids (list): The list of organisms found in the Mycocosm table.

    Returns:
        pd.DataFrame: The table data as a pandas DataFrame.
    """
    found = []
    if not os.path.exists(input_dir):
        logging.error(f"No {input_dir} folder found. Please run fetch_all_files() first.")
        return pd.DataFrame()

    for organism_id in organism_ids:
        json_files = [f for f in os.listdir(input_dir) if f.startswith(f"all_files_{organism_id}_page_") and f.endswith(".json")]
        if not json_files:
            logging.warning(f"No JSON files found for {organism_id} in {input_dir}. Skipping.")
            found.append({
                "organism": organism_id,
                "file_name": "NO FILES FOUND",
                "file_id": "",
                "_id": "",
                "file_status": "",
                "md5sum": "",
                "file_date": "",
                "ncbi_taxon_id": "",
                "jat_label": "",
                "ncbi_taxon_class": "",
                "ncbi_taxon_family": "",
                "ncbi_taxon_order": "",
                "ncbi_taxon_genus": "",
                "ncbi_taxon_species": "",
                "file_type": "",
                "portal_display_location": ""
            })
            continue

        for json_file in sorted(json_files):
            json_file_path = os.path.join(input_dir, json_file)
            with open(json_file_path, "r") as f:
                data = json.load(f)

            files = data.get("organisms", [{}])[0].get("files", [])
            for file in files:
                metadata = file.get("metadata", {})
                ncbi_taxon = metadata.get("ncbi_taxon", {})
                portal = metadata.get("portal", {})

                found.append({
                    "organism": organism_id,
                    "file_name": file.get("file_name"),
                    "file_id": file.get("file_id"),
                    "_id": file.get("_id"),
                    "file_status": file.get("file_status"),
                    "md5sum": file.get("md5sum"),
                    "file_date": file.get("file_date"),
                    "ncbi_taxon_id": metadata.get("ncbi_taxon_id", ""),
                    "jat_label": metadata.get("jat_label", ""),
                    "ncbi_taxon_class": ncbi_taxon.get("ncbi_taxon_class", ""),
                    "ncbi_taxon_family": ncbi_taxon.get("ncbi_taxon_family", ""),
                    "ncbi_taxon_order": ncbi_taxon.get("ncbi_taxon_order", ""),
                    "ncbi_taxon_genus": ncbi_taxon.get("ncbi_taxon_genus", ""),
                    "ncbi_taxon_species": ncbi_taxon.get("ncbi_taxon_species", ""),
                    "file_type": file.get("file_type", ""),
                    "portal_display_location": portal.get("display_location", "")
                })

    df = pd.DataFrame(found)
    logging.info(f"Parsed metadata for {len(df)} files.")
    return df